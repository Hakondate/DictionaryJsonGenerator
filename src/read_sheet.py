import re

from openpyxl.worksheet.worksheet import Worksheet

def read_sheet(sheet: Worksheet):
    foods = []

    for i in range(13, sheet.max_row):
        group = int(convert_to_double(sheet.cell(row = i, column = 2).value) // 1000)
        name = sheet.cell(row = i, column = 4).value
        energy = convert_to_double(sheet.cell(row = i, column = 7).value)
        protein = convert_to_double(sheet.cell(row = i, column = 10).value)
        lipid = convert_to_double(sheet.cell(row = i, column = 13).value)
        carbohydrate = convert_to_double(sheet.cell(row = i, column = 21).value)
        sodium = convert_to_double(sheet.cell(row = i, column = 24).value)
        calcium = convert_to_double(sheet.cell(row = i, column = 26).value)
        magnesium = convert_to_double(sheet.cell(row = i, column = 27).value)
        iron = convert_to_double(sheet.cell(row = i, column = 29).value)
        zinc = convert_to_double(sheet.cell(row = i, column = 30).value)
        retinol = convert_to_double(sheet.cell(row = i, column = 38).value)
        vitaminB1 = convert_to_double(sheet.cell(row = i, column = 50).value)
        vitaminB2 = convert_to_double(sheet.cell(row = i, column = 51).value)
        vitaminC = convert_to_double(sheet.cell(row = i, column = 59).value)
        dietaryFiber = convert_to_double(sheet.cell(row = i, column = 19).value)
        salt = convert_to_double(sheet.cell(row = i, column = 61).value)
        note = sheet.cell(row = i, column = 62).value

        foods.append({
            'group': group,
            'name': name,
            'energy': energy,
            'protein': protein,
            'lipid': lipid,
            'carbohydrate': carbohydrate,
            'sodium': sodium,
            'calcium': calcium,
            'magnesium': magnesium,
            'iron': iron,
            'zinc': zinc,
            'retinol': retinol,
            'vitaminB1': vitaminB1,
            'vitaminB2': vitaminB2,
            'vitaminC': vitaminC,
            'dietaryFiber': dietaryFiber,
            'salt': salt,
            'note': note,
        })
    
    return {
        'dictionary': foods,
    }

def convert_to_double(input):
    if (type(input) is float or type(input) is int):
        return float(input)
    
    if (type(input) is not str):
        return 0
    
    matchs = re.search(r'\d+', input)
    if (matchs is not None and len(matchs.regs) > 0):
        return float(matchs.group())
    
    return 0
